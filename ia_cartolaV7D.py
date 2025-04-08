import requests
import pandas as pd
import numpy as np
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Variável global para o DataFrame
df = None

def main():
    global df
    
    # 1. Interface do usuário
    saldo, num_times, estrategia = obter_configuracoes_usuario()
    
    # 2. Coletar dados do mercado
    print("\nColetando dados do Cartola FC...")
    dados_mercado = fetch_market_data()
    if not dados_mercado:
        exit("Erro: Não foi possível conectar à API do Cartola FC. Verifique sua internet.")
    
    # 3. Processar dados
    print("Processando dados dos jogadores...")
    df = processar_dados(dados_mercado)
    df = calcular_score_por_posicao(df)
    
    # 4. Gerar times
    print(f"\nGerando {num_times} time(s) com estratégia '{estrategia}'...")
    times = []
    saldo_restante = saldo
    
    for i in range(num_times):
        print(f"Gerando time {i+1}...")
        time = criar_time_otimizado(df, saldo_restante, estrategia)
        if not time:
            print("Não foi possível gerar mais times com o saldo restante.")
            break
        
        custo_time = calcular_custo_time(time)
        saldo_restante -= custo_time
        times.append(time)
        print(f"Time {i+1} gerado! Custo: C$ {custo_time:.2f} | Saldo restante: C$ {saldo_restante:.2f}")
    
    if not times:
        exit("Erro: Não foi possível gerar nenhum time com o saldo disponível.")
    
    # 5. Mostrar resultados e exportar
    mostrar_resumo(times, saldo)
    arquivo = exportar_para_excel(times, df)
    print(f"\nTimes exportados para o arquivo: {arquivo}")
    print("Boa sorte no Cartola FC!")

def fetch_market_data():
    """Coleta dados do mercado da API do Cartola com tratamento de erros"""
    try:
        response = requests.get("https://api.cartolafc.globo.com/atletas/mercado", timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Erro ao acessar a API: {e}")
        return None

def processar_dados(dados_mercado):
    """Processa os dados brutos da API para um DataFrame"""
    jogadores = dados_mercado['atletas']
    clubes = dados_mercado['clubes']
    posicoes = dados_mercado['posicoes']
    
    dados = []
    for jogador in jogadores:
        clube_id = jogador.get('clube_id')
        clube_nome = clubes.get(str(clube_id), {}).get('nome', 'Desconhecido')
        posicao_id = jogador.get('posicao_id')
        posicao_nome = posicoes.get(str(posicao_id), {}).get('nome', 'Desconhecido')
        
        scout = jogador.get('scout', {})
        dados.append({
            "Apelido": jogador.get('apelido', 'Desconhecido'),
            "Clube": clube_nome,
            "Posição": posicao_nome,
            "Preço (C$)": jogador.get('preco_num', 0),
            "Média Pontuação": jogador.get('media_num', 0),
            "Última Pontuação": jogador.get('pontos_num', 0),
            "Mínimo para Valorizar": jogador.get('minimo_para_valorizar', 0),
            "Variação de Preço": jogador.get('variacao_num', 0),
            "Desarmes (DS)": scout.get('DS', 0),
            "Finalizações Defendidas (FD)": scout.get('FD', 0),
            "Finalizações Fora (FF)": scout.get('FF', 0),
            "Faltas Cometidas (FC)": scout.get('FC', 0),
            "Gols (G)": scout.get('G', 0),
            "Assistências (A)": scout.get('A', 0),
            "Finalizações na Trave (FT)": scout.get('FT', 0),
            "Impedimentos (I)": scout.get('I', 0),
            "Defesas (DE)": scout.get('DE', 0),
            "Status": None  # Será preenchido depois
        })
    
    return pd.DataFrame(dados)

def calcular_score_por_posicao(df):
    """Calcula um score personalizado para cada jogador baseado em sua posição"""
    pesos = {
        'Goleiro': {'media': 0.7, 'DS': 0.25, 'FD': 0.05, 'variacao': 0.0, 'DE': 0.3},
        'Zagueiro': {'media': 0.6, 'DS': 0.3, 'FD': 0.05, 'variacao': 0.05, 'G': 0.1},
        'Lateral': {'media': 0.55, 'DS': 0.2, 'FD': 0.05, 'variacao': 0.2, 'A': 0.15},
        'Meia': {'media': 0.6, 'DS': 0.1, 'FD': 0.05, 'variacao': 0.25, 'A': 0.2, 'G': 0.15},
        'Atacante': {'media': 0.65, 'DS': 0.0, 'FD': 0.1, 'variacao': 0.25, 'G': 0.3, 'FT': 0.1},
        'Técnico': {'media': 0.8, 'DS': 0.0, 'FD': 0.0, 'variacao': 0.2}
    }
    
    # Classificar status dos jogadores
    conditions = [
        (df['Média Pontuação'] > 7) & (df['Variação de Preço'] > 0),
        (df['Média Pontuação'] > 5) & (df['Variação de Preço'] >= 0),
        (df['Média Pontuação'] < 4) & (df['Variação de Preço'] < 0)
    ]
    choices = ['Ótimo', 'Bom', 'Risco']
    df['Status'] = np.select(conditions, choices, default='Regular')
    
    # Calcular score para cada posição
    def calculate_score(row):
        weight = pesos.get(row['Posição'], pesos['Meia'])
        score = (
            row['Média Pontuação'] * weight['media'] +
            row['Desarmes (DS)'] * weight['DS'] +
            row['Finalizações Defendidas (FD)'] * weight['FD'] +
            row['Variação de Preço'] * weight['variacao']
        )
        
        # Adicionar fatores específicos por posição
        if row['Posição'] == 'Goleiro':
            score += row['Defesas (DE)'] * weight['DE']
        elif row['Posição'] in ['Zagueiro', 'Meia', 'Atacante']:
            score += row['Gols (G)'] * weight.get('G', 0)
        if row['Posição'] in ['Lateral', 'Meia']:
            score += row['Assistências (A)'] * weight.get('A', 0)
        if row['Posição'] == 'Atacante':
            score += row['Finalizações na Trave (FT)'] * weight.get('FT', 0)
        
        return round(score, 2)
    
    df['Score'] = df.apply(calculate_score, axis=1)
    return df.sort_values('Score', ascending=False)

def criar_time_otimizado(df, saldo, estrategia='pontuacao', tentativas_max=500):
    """Algoritmo de seleção de time que otimiza conforme a estratégia escolhida"""
    posicoes_necessarias = {
        'Goleiro': 1,
        'Zagueiro': 2,
        'Lateral': 2,
        'Meia': 3,
        'Atacante': 3,
        'Técnico': 1
    }
    
    # Filtrar top jogadores por posição
    top_jogadores = {
        pos: df[df['Posição'] == pos].sort_values('Score', ascending=False).head(30).to_dict('records')
        for pos in posicoes_necessarias.keys()
    }
    
    melhor_time = None
    melhor_metrica = -1
    
    for _ in range(tentativas_max):
        time = {}
        custo_total = 0
        metrica = 0
        
        for pos, qtd in posicoes_necessarias.items():
            jogadores_pos = top_jogadores[pos]
            
            if qtd == 1:
                jogador = random.choice(jogadores_pos)
                time[pos] = jogador
                custo_total += jogador['Preço (C$)']
                
                if estrategia == 'pontuacao':
                    metrica += jogador['Score']
                elif estrategia == 'valorizacao':
                    metrica += jogador['Variação de Preço']
                else:  # equilibrado
                    metrica += jogador['Score'] * 0.7 + jogador['Variação de Preço'] * 0.3
            else:
                selecionados = random.sample(jogadores_pos, qtd)
                time[pos + 's'] = selecionados
                custo_total += sum(j['Preço (C$)'] for j in selecionados)
                
                if estrategia == 'pontuacao':
                    metrica += sum(j['Score'] for j in selecionados)
                elif estrategia == 'valorizacao':
                    metrica += sum(j['Variação de Preço'] for j in selecionados)
                else:
                    metrica += sum(j['Score'] * 0.7 + j['Variação de Preço'] * 0.3 for j in selecionados)
        
        if custo_total <= saldo and metrica > melhor_metrica:
            melhor_time = time
            melhor_metrica = metrica
    
    return melhor_time if melhor_time else criar_time_fallback(df, saldo)

def criar_time_fallback(df, saldo):
    """Método de fallback mais simples para quando o otimizado falha"""
    posicoes = {
        'Goleiro': 1,
        'Zagueiro': 2,
        'Lateral': 2,
        'Meia': 3,
        'Atacante': 3,
        'Técnico': 1
    }
    
    for _ in range(100):  # Tentativas máximas
        time = {}
        custo_total = 0
        
        for pos, qtd in posicoes.items():
            if qtd == 1:
                jogador = random.choice(df[df['Posição'] == pos].head(10).to_dict('records'))
                time[pos] = jogador
                custo_total += jogador['Preço (C$)']
            else:
                jogadores = random.sample(df[df['Posição'] == pos].head(15).to_dict('records'), qtd)
                time[pos + 's'] = jogadores
                custo_total += sum(j['Preço (C$)'] for j in jogadores)
        
        if custo_total <= saldo:
            return time
    
    return None

def calcular_custo_time(time):
    """Calcula o custo total de um time"""
    custo = 0
    
    # Verificar cada posição no time
    if 'Goleiro' in time:
        custo += time['Goleiro']['Preço (C$)']
    
    if 'Zagueiros' in time:
        custo += sum(j['Preço (C$)'] for j in time['Zagueiros'])
    
    if 'Laterais' in time:
        custo += sum(j['Preço (C$)'] for j in time['Laterais'])
    
    if 'Meias' in time:
        custo += sum(j['Preço (C$)'] for j in time['Meias'])
    
    if 'Atacantes' in time:
        custo += sum(j['Preço (C$)'] for j in time['Atacantes'])
    
    if 'Técnico' in time:
        custo += time['Técnico']['Preço (C$)']
    
    return custo

def obter_configuracoes_usuario():
    """Interface para obter as configurações do usuário"""
    print("\n" + "="*50)
    print(" IA DE RECOMENDAÇÃO PARA CARTOLA FC ".center(50, "="))
    print("="*50 + "\n")
    
    while True:
        try:
            saldo = float(input("• Digite seu saldo disponível em cartoletas: C$ "))
            if saldo <= 0:
                print("Valor deve ser positivo. Tente novamente.")
                continue
                
            num_times = int(input("• Quantos times deseja gerar (1-10)? "))
            if num_times < 1 or num_times > 10:
                print("Por favor, digite um número entre 1 e 10.")
                continue
                
            print("\nEscolha a estratégia principal:")
            print("1. Máxima Pontuação (melhores jogadores)")
            print("2. Valorização (jogadores que mais valorizam)")
            print("3. Equilibrado (balanceado entre pontuação e valorização)")
            opcao = input("• Digite o número da estratégia desejada (1-3): ")
            
            if opcao == '1':
                estrategia = 'pontuacao'
            elif opcao == '2':
                estrategia = 'valorizacao'
            elif opcao == '3':
                estrategia = 'equilibrado'
            else:
                print("Opção inválida. Usando estratégia equilibrada.")
                estrategia = 'equilibrado'
            
            print("\n" + "="*50)
            print(f" Configuração: {num_times} time(s) | C$ {saldo:.2f} | Estratégia: {estrategia.capitalize()} ".center(50, "="))
            print("="*50 + "\n")
            
            confirmacao = input("Confirmar configurações (S/N)? ").lower()
            if confirmacao == 's':
                return saldo, num_times, estrategia
            else:
                print("\nVamos começar novamente...\n")
                
        except ValueError:
            print("Entrada inválida. Por favor, digite um número.")
            continue

def exportar_para_excel(times, df_completo, filename="times_recomendados_avancado.xlsx"):
    """Exporta os times para um arquivo Excel com formatação profissional"""
    # Preparar dados dos times
    dados_saida = []
    for i, time in enumerate(times, 1):
        todos_jogadores = []
        for pos, jgs in time.items():
            if isinstance(jgs, list):
                todos_jogadores.extend(jgs)
            else:
                todos_jogadores.append(jgs)
        capitao = max(todos_jogadores, key=lambda x: x["Média Pontuação"])
        
        for posicao, jogadores in time.items():
            if isinstance(jogadores, list):
                for jogador in jogadores:
                    is_capitao = jogador["Apelido"] == capitao["Apelido"]
                    dados_saida.append([
                        f"Time {i}", posicao, 
                        jogador["Apelido"] + (" (Capitão)" if is_capitao else ""),
                        jogador["Clube"],
                        jogador["Posição"],
                        jogador["Preço (C$)"],
                        jogador["Média Pontuação"],
                        jogador["Última Pontuação"],
                        jogador["Variação de Preço"],
                        jogador["Status"]
                    ])
            else:
                is_capitao = jogadores["Apelido"] == capitao["Apelido"]
                dados_saida.append([
                    f"Time {i}", posicao, 
                    jogadores["Apelido"] + (" (Capitão)" if is_capitao else ""),
                    jogadores["Clube"],
                    jogadores["Posição"],
                    jogadores["Preço (C$)"],
                    jogadores["Média Pontuação"],
                    jogadores["Última Pontuação"],
                    jogadores["Variação de Preço"],
                    jogadores["Status"]
                ])
    
    # Criar DataFrames para exportação
    cols = ["Time", "Posição", "Jogador", "Clube", "Posição", "Preço", "Média", "Última", "Variação", "Status"]
    saida_df = pd.DataFrame(dados_saida, columns=cols)
    
    # Criar Excel com múltiplas abas
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Aba com os times
        saida_df.to_excel(writer, sheet_name='Times Recomendados', index=False)
        
        # Aba com estatísticas
        estatisticas = []
        for i, time in enumerate(times, 1):
            todos_jogadores = []
            for pos, jgs in time.items():
                if isinstance(jgs, list):
                    todos_jogadores.extend(jgs)
                else:
                    todos_jogadores.append(jgs)
            
            estatisticas.append([
                f"Time {i}",
                sum(j["Preço (C$)"] for j in todos_jogadores),
                sum(j["Média Pontuação"] for j in todos_jogadores),
                sum(j["Variação de Preço"] for j in todos_jogadores),
                len([j for j in todos_jogadores if j["Status"] == "Ótimo"]),
                len([j for j in todos_jogadores if j["Status"] == "Risco"])
            ])
        
        stats_df = pd.DataFrame(estatisticas, 
                              columns=["Time", "Custo Total", "Média Total", "Valorização Total", 
                                       "Jogadores Ótimos", "Jogadores de Risco"])
        stats_df.to_excel(writer, sheet_name='Estatísticas', index=False)
        
        # Aba com melhores por posição
        for pos in df_completo['Posição'].unique():
            (df_completo[df_completo['Posição'] == pos]
             .sort_values('Score', ascending=False)
             .head(15)
             .to_excel(writer, sheet_name=f'Top {pos}s', index=False))
        
        # Aba com todos os jogadores
        df_completo.sort_values('Score', ascending=False).to_excel(
            writer, sheet_name='Todos Jogadores', index=False)
    
    # Aplicar formatação condicional
    wb = load_workbook(filename)
    
    # Cores para status
    status_colors = {
        'Ótimo': 'FF00B050',  # Verde
        'Bom': 'FF92D050',     # Verde claro
        'Regular': 'FFFFC000', # Amarelo
        'Risco': 'FFC00000'    # Vermelho
    }
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Formatar cabeçalho
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0070C0")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatar coluna de status (se existir)
        if sheet == 'Times Recomendados':
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter == 'J':  # Coluna J = Status
                        status = cell.value
                        if status in status_colors:
                            cell.fill = PatternFill("solid", fgColor=status_colors[status])
        
        # Ajustar largura das colunas
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
    
    wb.save(filename)
    return filename

def mostrar_resumo(times, saldo):
    """Exibe um resumo detalhado dos times gerados"""
    global df
    
    print("\n" + "="*50)
    print(" RESUMO DOS TIMES GERADOS ".center(50, "="))
    print("="*50)
    
    for i, time in enumerate(times, 1):
        todos_jogadores = []
        for pos, jgs in time.items():
            if isinstance(jgs, list):
                todos_jogadores.extend(jgs)
            else:
                todos_jogadores.append(jgs)
        
        custo = sum(j["Preço (C$)"] for j in todos_jogadores)
        media = sum(j["Média Pontuação"] for j in todos_jogadores)
        valorizacao = sum(j["Variação de Preço"] for j in todos_jogadores)
        capitao = max(todos_jogadores, key=lambda x: x["Média Pontuação"])
        
        print(f"\n• Time {i}:")
        print(f"  - Custo total: C$ {custo:.2f} (Saldo restante: C$ {saldo - custo:.2f})")
        print(f"  - Média de pontuação: {media:.1f}")
        print(f"  - Valorização média: {valorizacao:.2f}")
        print(f"  - Capitão: {capitao['Apelido']} ({capitao['Média Pontuação']} pts)")
        
        # Contar status
        status_counts = {'Ótimo': 0, 'Bom': 0, 'Regular': 0, 'Risco': 0}
        for j in todos_jogadores:
            status_counts[j.get('Status', 'Regular')] += 1
        
        print("  - Distribuição de status:")
        for status, count in status_counts.items():
            if count > 0:
                print(f"    {status}: {count} jogador(es)")
    
    print("\n" + "="*50)
    print(" RECOMENDAÇÕES ".center(50, "="))
    print("="*50)
    
    # Mostrar destaques
    print("\n• Top 3 jogadores com maior potencial de pontuação:")
    print(df.sort_values('Score', ascending=False)[['Apelido', 'Clube', 'Posição', 'Score']].head(3).to_string(index=False))
    
    print("\n• Top 3 jogadores com maior valorização recente:")
    print(df.sort_values('Variação de Preço', ascending=False)[['Apelido', 'Clube', 'Posição', 'Variação de Preço']].head(3).to_string(index=False))
    
    print("\n• 3 jogadores com melhor custo-benefício:")
    df['Custo_Beneficio'] = df['Score'] / df['Preço (C$)']
    print(df.sort_values('Custo_Beneficio', ascending=False)[['Apelido', 'Clube', 'Posição', 'Custo_Beneficio']].head(3).to_string(index=False))

if __name__ == "__main__":
    main()